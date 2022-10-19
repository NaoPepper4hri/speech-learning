#!/usr/bin/python
"""
Generate a JSON description of the questions in the experiment.

Given their original excel description, we can create a set of questions in JSON format that can be
loaded directly from the webapp.
"""
import argparse
import json
import openpyxl as xml
from typing import Any, Dict, Iterator, Optional, Tuple


class DataSheets:
    """Name of the excel sheets corresponding to each question type."""

    IMAGE_BUTTON_SHEET = "Block 1"
    FILL_BLANK_SHEET = "Block 2"
    VOCABULARY_SHEET = "Block 3"
    SORT_WORDS_SHEET = "Block 4"
    SORT_WORDS_SHEET_2 = "Block 4 (Difficult Version)"
    MATCH_PAIRS_SHEET = "Block 5"


class Question:
    """Base description of a question."""

    def __init__(self, id: Any, question: Any, options: Any) -> None:
        """Initialize the class, given the question id, content and other options."""
        self.id = id
        self.question = question
        self.options = options

    def as_dict(self) -> Dict:
        """Convert this question into a dictionary (defines the shape of the JSON written.)."""
        return {
            "ty": self.__class__.__name__,
            "id": self.id,
            "question": self.question,
            "options": self.options,
        }


class ImageButtonQuestion(Question):
    """Question including a set of buttons with images."""

    @classmethod
    def from_row(cls, id: str, data: Tuple):
        """Create a Question from a excel row."""
        images, question, answer, options = data[0:4]
        if images is None:
            images = ",,"
        opt = []
        for image, text in zip(images.split(","), options.split(",")):
            image = image.strip()
            text = text.strip()
            opt.append({"text": text, "image": image, "correct": text == answer})
        return cls(id=id, question=question, options=opt)


class FillBlankQuestion(Question):
    """A sentence with a missing word.Several options are given, but only one is correct."""

    def __init__(self, image, translation, *args, **kwargs) -> None:
        """Create a question, adding the specific fields for this type."""
        super().__init__(*args, **kwargs)
        self.image = image
        self.translation = translation

    def as_dict(self) -> Dict:
        """Convert this question into a dictionary (defines the shape of the JSON written.)."""
        base = super().as_dict()
        base["image"] = self.image
        base["translation"] = self.translation
        return base

    @classmethod
    def from_row(cls, id: Any, data: Tuple):
        """Create a Question from a excel row."""
        image, question, answer, options, translation = data[0:5]
        opt = {}
        counter = 0
        for o in options.split(","):
            key = "o{}".format(counter)
            counter += 1
            o = o.strip()
            opt[key] = {"id": key, "text": o, "correct": o == answer}
        return cls(id=id, image=image, translation=translation, question=question, options=opt)


class VocabularyQuestion(Question):
    """A selection of words and images from which a translation should be selected."""

    @classmethod
    def from_row(cls, id: Any, data: Tuple):
        """Create a Question from a excel row."""
        print(data)
        question, answer, options = data[0:3]
        opts = []
        for o in options.split(","):
            o = o.strip()
            opts.append({"text": o, "correct": o == answer})
        return cls(
            id=id,
            question=question,
            options=opts,
        )


class SortWordsQuestion(Question):
    """Given a sentence in language A, sort the options in language B to find its translation."""

    def __init__(self, header, answer, *args, **kwargs) -> None:
        """Create a question, adding the specific fields for this type."""
        super().__init__(*args, **kwargs)
        self.header = header
        self.answer = answer

    def as_dict(self) -> Dict:
        """Convert this question into a dictionary (defines the shape of the JSON written.)."""
        base = super().as_dict()
        base["answer"] = self.answer
        base["header"] = self.header
        return base

    @classmethod
    def from_row(cls, id: Any, data: Tuple, task: Optional[str] = ""):
        """Create a Question from a excel row."""
        question, answer, options = data[0:3]
        opts = {}
        counter = 0
        answer_sequence = answer.split()
        print(answer)
        for o in options.split(","):
            o = o.strip()
            key = "o{}".format(counter)
            counter += 1
            opts[key] = {
                "id": key,
                "tIdx": answer_sequence.index(o) if o in answer_sequence else -1,
                "text": o,
            }
        return cls(
            header=task,
            answer=answer,
            id=id,
            question=question,
            options=opts,
        )


class MatchPairsQuestion(Question):
    """Present a set of word pairs (e.g. `arbol` - `tree`)."""

    @classmethod
    def from_sheet(cls, question: str, data: Iterator):
        """Create a Question from a excel sheet."""
        opts = []
        for row in data:
            opts.append(
                {
                    "p1": row[0],
                    "p2": row[1],
                }
            )

        return cls(
            id="pairs",
            question=question,
            options=opts,
        )


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("data_file", type=str, help="Source excel file.")
    parser.add_argument("output", type=str, help="Destination js file.")

    args = parser.parse_args()

    if args.data_file:
        book = xml.load_workbook(args.data_file)
        questions = []
        pairs = None

        for name in book.sheetnames:
            question_type = name
            sheet = book[name]
            task = sheet["A2"].value
            if question_type == DataSheets.MATCH_PAIRS_SHEET:
                pairs = MatchPairsQuestion.from_sheet(
                    task, sheet.iter_rows(min_row=3, values_only=True)
                )
            else:
                idx = 0
                for entry in sheet.iter_rows(min_row=3, values_only=True):
                    idx += 1
                    match question_type:
                        case DataSheets.IMAGE_BUTTON_SHEET:
                            questions.append(
                                ImageButtonQuestion.from_row("{}_{}".format(name, idx), entry)
                            )
                        case DataSheets.FILL_BLANK_SHEET:
                            questions.append(
                                FillBlankQuestion.from_row("{}_{}".format(name, idx), entry)
                            )
                        case DataSheets.VOCABULARY_SHEET:
                            questions.append(
                                VocabularyQuestion.from_row("{}_{}".format(name, idx), entry)
                            )
                        case DataSheets.SORT_WORDS_SHEET:
                            try:
                                questions.append(
                                    SortWordsQuestion.from_row(
                                        "{}_{}".format(name, idx), entry, task
                                    )
                                )
                            except Exception as e:
                                print(e)
                        case DataSheets.SORT_WORDS_SHEET_2:
                            try:
                                questions.append(
                                    SortWordsQuestion.from_row(
                                        "{}_{}".format(name, idx), entry, task
                                    )
                                )
                            except Exception as e:
                                print(e)

        json_obj = [q.as_dict() for q in questions]
        if pairs is not None:
            json_obj.append(pairs.as_dict())

        with open(args.output, "w", encoding="utf-8") as f:
            json.dump(json_obj, f, ensure_ascii=False, indent=4)
