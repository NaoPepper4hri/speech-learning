#!/usr/bin/python
import argparse
import json
import openpyxl as xml
from typing import Dict, Iterator, Optional, Tuple


class DataSheets:
    IMAGE_BUTTON_SHEET = "Block 1"
    FILL_BLANK_SHEET = "Block 2"
    VOCABULARY_SHEET = "Block 3"
    SORT_WORDS_SHEET = "Block 4"
    SORT_WORDS_SHEET_2 = "Block 4 (Difficult Version)"
    MATCH_PAIRS_SHEET = "Block 5"


class Question:
    def __init__(self, question, options) -> None:
        self.question = question
        self.options = options

    def as_dict(self) -> Dict:
        return {
            "ty": self.__class__.__name__,
            "question": self.question,
            "options": self.options,
        }


class ImageButtonQuestion(Question):
    @classmethod
    def from_row(cls, data: Tuple):
        images, question, answer, options = data
        if images is None:
            images = ",,"
        opt = []
        for image, text in zip(images.split(","), options.split(",")):
            image = image.strip()
            text = text.strip()
            opt.append({
                "text": text,
                "image": image,
                "correct": text == answer
            })
        return cls(
            question=question,
            options=opt
        )


class FillBlankQuestion(Question):
    def __init__(self, image, translation, *args, **kwargs) -> None:
        super().__init__(*args, **kwargs)
        self.image = image
        self.translation = translation

    def as_dict(self) -> Dict:
        return {
            "ty": self.__class__.__name__,
            "question": self.question,
            "image": self.image,
            "options": self.options,
            "translation": self.translation
        }

    @classmethod
    def from_row(cls, data: Tuple):
        image, question, answer, options, translation = data
        opt = {}
        counter = 0
        for o in options.split(","):
            key = "o{}".format(counter)
            counter += 1
            o = o.strip()
            opt[key] = {
                    "id": key,
                    "text": o,
                    "correct": o == answer
            }
        return cls(
            image=image,
            translation=translation,
            question=question,
            options=opt
        )


class VocabularyQuestion(Question):
    @classmethod
    def from_row(cls, data: Tuple):
        question, answer, options = data
        opts = []
        for o in options.split(","):
            o = o.strip()
            opts.append({
                "text": o,
                "correct": o == answer
            })
        return cls(
            question=question,
            options=opts,
        )


class SortWordsQuestion(Question):
    def __init__(self, header, answer, *args, **kwargs) -> None:
        super().__init__(*args, **kwargs)
        self.header = header
        self.answer = answer

    def as_dict(self) -> Dict:
        return {
            "ty": self.__class__.__name__,
            "question": self.question,
            "answer": self.answer,
            "header": self.header,
            "options": self.options,
        }

    @classmethod
    def from_row(cls, data: Tuple, task: Optional[str] = ""):
        question, answer, options = data[0:3]
        opts = {}
        counter = 0
        answer_sequence = answer.split()
        print(answer)
        for o in options.split(","):
            o = o.strip()
            key = "o{}".format(counter)
            counter += 1
            opts[key] = {
                "id": key,
                "tIdx": answer_sequence.index(o) if o in answer_sequence else -1,
                "text": o,
            }
        return cls(
            header=task,
            answer=answer,
            question=question,
            options=opts,
        )


class MatchPairsQuestion(Question):
    @classmethod
    def from_sheet(cls, question: str, data: Iterator):
        opts = []
        for row in data:
            opts.append({
                "p1": row[0],
                "p2": row[1],
                })

        return cls(
            question=question,
            options=opts,
        )


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("data_file", type=str, help="Source excel file.")
    parser.add_argument("output", type=str, help="Destination js file.")

    args = parser.parse_args()

    if args.data_file:
        book = xml.load_workbook(args.data_file)
        questions = []
        pairs = None

        for name in book.sheetnames:
            question_type = name
            sheet = book[name]
            task = sheet["A2"].value
            if question_type == DataSheets.MATCH_PAIRS_SHEET:
                pairs = MatchPairsQuestion.from_sheet(
                        task,
                        sheet.iter_rows(min_row=3, values_only=True))
            else:
                for entry in sheet.iter_rows(min_row=3, values_only=True):
                    match question_type:
                        case DataSheets.IMAGE_BUTTON_SHEET:
                            questions.append(ImageButtonQuestion.from_row(entry))
                        case DataSheets.FILL_BLANK_SHEET:
                            questions.append(FillBlankQuestion.from_row(entry))
                        case DataSheets.VOCABULARY_SHEET:
                            questions.append(VocabularyQuestion.from_row(entry))
                        case DataSheets.SORT_WORDS_SHEET:
                            try:
                                questions.append(SortWordsQuestion.from_row(entry, task))
                            except Exception as e:
                                print(e)
                        case DataSheets.SORT_WORDS_SHEET_2:
                            try:
                                questions.append(SortWordsQuestion.from_row(entry, task))
                            except Exception as e:
                                print(e)

        json_obj = [q.as_dict() for q in questions]
        json_obj.append(pairs.as_dict())

        with open(args.output, 'w', encoding='utf-8') as f:
            json.dump(json_obj, f, ensure_ascii=False, indent=4)
